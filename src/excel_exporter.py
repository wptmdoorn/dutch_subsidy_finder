"""
Excel Exporter for Dutch Subsidy Finder
Exports processed subsidy data to Excel format
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from .config import Config


class ExcelExporter:
    """Exports subsidy data to Excel with formatting."""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
    def export_subsidies(self, subsidies: List[Dict]) -> str:
        """Export subsidies to Excel file."""
        
        # Ensure output directory exists
        Config.OUTPUT_DIR.mkdir(exist_ok=True)
        
        # Create output filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"dutch_subsidies_{timestamp}.xlsx"
        output_path = Config.OUTPUT_DIR / filename
        
        self.logger.info(f"Exporting {len(subsidies)} subsidies to {output_path}")
        
        # Create DataFrame
        df = self._create_dataframe(subsidies)
        
        # Create Excel workbook with formatting
        self._create_formatted_excel(df, output_path)
        
        self.logger.info(f"Successfully exported to {output_path}")
        return str(output_path)
    
    def _create_dataframe(self, subsidies: List[Dict]) -> pd.DataFrame:
        """Create pandas DataFrame from subsidies data."""
        
        # Prepare data for DataFrame
        data = []
        
        for subsidy in subsidies:
            row = {
                'Subsidy Name': subsidy.get('name', ''),
                'Funding Organization': subsidy.get('funding_organization', ''),
                'Amount/Budget': subsidy.get('amount', ''),
                'Application Deadline': subsidy.get('deadline', ''),
                'Status': subsidy.get('status', ''),
                'Eligibility Criteria': subsidy.get('eligibility', ''),
                'Research Areas': subsidy.get('research_areas', ''),
                'Description': subsidy.get('description', ''),
                'Application Process': subsidy.get('application_process', ''),
                'Contact Information': subsidy.get('contact_info', ''),
                'Website URL': subsidy.get('url', ''),
                'Relevance Score': subsidy.get('relevance_score', 0),
                'Keywords Matched': subsidy.get('keywords_matched', ''),
                'Researcher Level': subsidy.get('researcher_level', ''),
                'Date Scraped': subsidy.get('date_scraped', '')
            }
            data.append(row)
        
        df = pd.DataFrame(data)
        
        # Sort by relevance score (highest first)
        df = df.sort_values('Relevance Score', ascending=False)
        
        return df
    
    def _create_formatted_excel(self, df: pd.DataFrame, output_path: Path):
        """Create formatted Excel file."""
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Dutch Subsidies"
        
        # Add title (merge across all columns)
        last_col = chr(ord('A') + len(df.columns) - 1)
        ws.merge_cells(f'A1:{last_col}1')
        title_cell = ws['A1']
        title_cell.value = f"Dutch Research & Innovation Subsidies - Clinical Chemistry & AI Focus"
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add subtitle with generation info
        ws.merge_cells(f'A2:{last_col}2')
        subtitle_cell = ws['A2']
        subtitle_cell.value = f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total: {len(df)} subsidies"
        subtitle_cell.font = Font(size=12, italic=True)
        subtitle_cell.alignment = Alignment(horizontal="center")
        
        # Add data starting from row 4
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format headers (row 4)
        header_row = 4
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Freeze header row
        ws.freeze_panes = f'A5'
        
        # Auto-fit column widths (avoid merged cell errors)
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = min(adjusted_width, 50)
        
        # Format data rows
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply formatting to all data cells
        for row in range(header_row, ws.max_row + 1):
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                # Special formatting for specific columns
                if col == df.columns.get_loc('Relevance Score') + 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    # Color code relevance scores
                    score = cell.value
                    if isinstance(score, (int, float)):
                        if score >= 8:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
                        elif score >= 6:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
                        elif score >= 4:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
                
                elif col == df.columns.get_loc('Status') + 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if cell.value == "Open":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
                    elif cell.value == "Closed":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
                
                elif col == df.columns.get_loc('Website URL') + 1:
                    if cell.value:
                        cell.font = Font(color="0000FF", underline="single")  # Blue hyperlink style
        
        # Add summary sheet
        self._add_summary_sheet(wb, df)
        
        # Save workbook
        wb.save(output_path)
    
    def _add_summary_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Add summary statistics sheet."""
        
        ws_summary = wb.create_sheet("Summary")
        
        # Title
        ws_summary['A1'] = "Dutch Subsidies Summary Report"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary.merge_cells('A1:C1')
        
        # Basic statistics
        row = 3
        stats = [
            ("Total Subsidies Found:", len(df)),
            ("High Relevance (Score ≥8):", len(df[df['Relevance Score'] >= 8])),
            ("Medium Relevance (Score 6-7.9):", len(df[(df['Relevance Score'] >= 6) & (df['Relevance Score'] < 8)])),
            ("Open Status:", len(df[df['Status'] == 'Open'])),
            ("Closed Status:", len(df[df['Status'] == 'Closed'])),
            ("Average Relevance Score:", round(df['Relevance Score'].mean(), 2)),
        ]
        
        for label, value in stats:
            ws_summary[f'A{row}'] = label
            ws_summary[f'B{row}'] = value
            ws_summary[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Funding organizations breakdown
        row += 2
        ws_summary[f'A{row}'] = "By Funding Organization:"
        ws_summary[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        org_counts = df['Funding Organization'].value_counts()
        for org, count in org_counts.items():
            ws_summary[f'A{row}'] = org
            ws_summary[f'B{row}'] = count
            row += 1
        
        # Top keywords
        row += 2
        ws_summary[f'A{row}'] = "Most Common Keywords:"
        ws_summary[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        # Count keyword occurrences
        all_keywords = []
        for keywords_str in df['Keywords Matched'].dropna():
            if keywords_str:
                keywords = [k.strip() for k in keywords_str.split(',')]
                all_keywords.extend(keywords)
        
        from collections import Counter
        keyword_counts = Counter(all_keywords)
        
        for keyword, count in keyword_counts.most_common(10):
            ws_summary[f'A{row}'] = keyword
            ws_summary[f'B{row}'] = count
            row += 1
        
        # Set column widths
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 15
